"""leads.db 의 리드 + 접촉이력을 엑셀(.xlsx) 로 내보내기."""

from __future__ import annotations

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from db import get_leads, get_contacts, Lead
from paths import resource_path


HEADERS = [
    "ID", "사이트", "게시판", "카테고리", "제목", "회사명",
    "카톡ID", "오픈채팅", "전화", "이메일",
    "작성자", "게시일", "상태", "매칭키워드",
    "접촉횟수", "마지막접촉", "마지막결과", "메모",
    "본문요약", "본문전체", "글URL",
]


def export_leads(out_path: str | None = None, site: str | None = None,
                 status: str | None = None) -> str:
    """엑셀 파일 생성. out_path 안 주면 leads_YYYYMMDD_HHMM.xlsx 자동 이름."""
    if not out_path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        out_path = resource_path(f"leads_{stamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "leads"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for c, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    leads = get_leads(site=site, status=status, limit=10_000)
    for i, L in enumerate(leads, start=2):
        contacts = get_contacts(L.id) if L.id else []
        last = contacts[0] if contacts else None
        last_note = last.note if last else ""

        row = [
            L.id,
            L.site,
            L.board,
            L.category,
            L.title,
            L.company,
            ", ".join(L.kakao_ids),
            ", ".join(L.open_chats),
            ", ".join(L.phones),
            ", ".join(L.emails),
            L.writer,
            L.posted_at,
            L.status,
            ", ".join(L.matched_keywords),
            len(contacts),
            last.attempted_at if last else "",
            last.result if last else "",
            last_note,
            L.body_excerpt,
            L.body_text,
            L.post_url,
        ]
        for c, val in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=val)

    # 컬럼 너비 조정
    widths = [6, 10, 12, 14, 40, 18, 24, 32, 18, 28, 16, 16, 10, 18, 8, 18, 10, 28, 60, 80, 60]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)].width = w

    # 헤더 행 고정
    ws.freeze_panes = "A2"

    # 접촉 이력 시트
    ws2 = wb.create_sheet("contacts")
    ws2.append(["LeadID", "사이트", "글제목", "시도시각", "채널", "결과", "메모"])
    for c, _ in enumerate(ws2[1], start=1):
        ws2.cell(row=1, column=c).font = header_font
        ws2.cell(row=1, column=c).fill = header_fill

    row = 2
    for L in leads:
        if not L.id:
            continue
        for ct in get_contacts(L.id):
            ws2.append([L.id, L.site, L.title, ct.attempted_at, ct.channel, ct.result, ct.note or ""])
            row += 1
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = export_leads()
    print(f"내보내기 완료: {path}")
